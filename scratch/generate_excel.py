import os
import random
import datetime
import subprocess
import sys

# Ensure openpyxl is installed
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set up output path
output_dir = r"c:\Users\SANTHU\OneDrive\Desktop\Projects\indhu-portfolio\public"
os.makedirs(output_dir, exist_ok=True)
excel_path = os.path.join(output_dir, "retail_sales_telemetry.xlsx")

# Create workbook
wb = Workbook()

# Styling options
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo
fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
font_regular = Font(name="Segoe UI", size=10)
font_bold = Font(name="Segoe UI", size=10, bold=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB')
)

# Sheet 1: README
ws_readme = wb.active
ws_readme.title = "README"
ws_readme.views.sheetView[0].showGridLines = True

readme_data = [
    ["Retail Operations & Sales Telemetry Dataset", ""],
    ["", ""],
    ["Dataset Overview", "This dataset is synthesized to match the metrics displayed on the Business Analyst Portfolio Dashboard."],
    ["It contains operational metrics spanning retail sales, operations leakage, inventory safety stock levels, and transaction processing errors.", ""],
    ["", ""],
    ["Sheets in this Workbook:", ""],
    ["1. README", "This guidance page."],
    ["2. KPI_Summary", "Summary dashboard with native Excel formulas rolling up metrics from the transaction databases."],
    ["3. Sales_Transactions", "Raw transactional database of sales, including leakage value and error status (1,500 rows)."],
    ["4. Inventory_Metrics", "Stock level snapshots across different warehouse hubs, showing stockouts and lead time warnings (300 rows)."],
    ["", ""],
    ["Metrics & Columns Glossary:", ""],
    ["- Revenue", "Gross revenue from sales transactions."],
    ["- Operations Leakage Value", "Revenue lost due to process inefficiencies, logistics errors, or damaged goods during transit."],
    ["- Safety Stock Compliance", "The inventory stock level relative to safety stock thresholds. Below 100% indicates under-stock risk."],
    ["- Process Error Rate", "The percentage of transactions that encountered processing validation errors or claims denials."],
    ["- Category", "Retail Category: Electronics (ELEC), Apparel (APPS), Home & Living (HLIV)."],
    ["- Warehouse Location", "Storage hub: Denver Warehouse, Seattle Import Hub, Chicago Depot."]
]

# Write README data
for r_idx, row in enumerate(readme_data, 1):
    ws_readme.append(row)
    # style title
    if r_idx == 1:
        ws_readme.cell(row=1, column=1).font = Font(name="Segoe UI", size=16, bold=True, color="4F46E5")
    elif row[0] in ["Dataset Overview", "Sheets in this Workbook:", "Metrics & Columns Glossary:"]:
        ws_readme.cell(row=r_idx, column=1).font = Font(name="Segoe UI", size=12, bold=True, color="1F2937")

# Auto fit column widths for README
ws_readme.column_dimensions['A'].width = 32
ws_readme.column_dimensions['B'].width = 110


# Sheet 2: Sales_Transactions
ws_sales = wb.create_sheet("Sales_Transactions")
ws_sales.views.sheetView[0].showGridLines = True

sales_headers = ["Transaction_ID", "Date", "Category", "Revenue", "Operations_Leakage_Pct", "Operations_Leakage_Value", "Processing_Status", "Error_Occurred"]
ws_sales.append(sales_headers)

# Apply header formatting
for col_idx in range(1, len(sales_headers) + 1):
    cell = ws_sales.cell(row=1, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

# Generate Sales Data
start_date = datetime.date(2026, 1, 1)
categories = ["Electronics", "Apparel", "Home & Living"]

# Error rates matching dashboard (ELEC 3.1%, APPS 1.8%, HLIV 2.4% avg)
error_rates = {
    "Electronics": 0.031,
    "Apparel": 0.018,
    "Home & Living": 0.024
}

# Leakage percentages
leakage_avgs = {
    "Electronics": 0.065,
    "Apparel": 0.045,
    "Home & Living": 0.055
}

random.seed(42) # For reproducibility

num_sales_rows = 1500
for i in range(num_sales_rows):
    txn_id = f"TXN-{10000 + i}"
    # Date distributed over 240 days (Jan to Aug)
    date_val = start_date + datetime.timedelta(days=random.randint(0, 240))
    category = random.choice(categories)
    
    # Revenue matching simulator levels (ELEC high, APPS med, HLIV low)
    if category == "Electronics":
        revenue = round(random.uniform(500, 3500), 2)
    elif category == "Apparel":
        revenue = round(random.uniform(50, 450), 2)
    else:
        revenue = round(random.uniform(100, 850), 2)
        
    # Leakage
    leakage_pct = round(random.uniform(0.01, leakage_avgs[category] * 2), 4)
    leakage_val = round(revenue * leakage_pct, 2)
    
    # Error occurred based on probabilities
    error_occurred = "Yes" if random.random() < error_rates[category] else "No"
    
    if error_occurred == "Yes":
        status = random.choice(["Denial", "Audit Review"])
    else:
        status = "Completed"
        
    ws_sales.append([txn_id, date_val, category, revenue, leakage_pct, leakage_val, status, error_occurred])

# Style Sales Data
for r_idx in range(2, num_sales_rows + 2):
    is_even = (r_idx % 2 == 0)
    for c_idx in range(1, len(sales_headers) + 1):
        cell = ws_sales.cell(row=r_idx, column=c_idx)
        cell.font = font_regular
        cell.border = thin_border
        if is_even:
            cell.fill = fill_zebra
            
        # Alignments & formats
        if c_idx in [1, 3, 7, 8]:
            cell.alignment = align_center if c_idx != 3 else align_left
        elif c_idx == 2:
            cell.alignment = align_center
            cell.number_format = 'yyyy-mm-dd'
        elif c_idx == 4:
            cell.alignment = align_right
            cell.number_format = '$#,##0.00'
        elif c_idx == 5:
            cell.alignment = align_right
            cell.number_format = '0.00%'
        elif c_idx == 6:
            cell.alignment = align_right
            cell.number_format = '$#,##0.00'


# Sheet 3: Inventory_Metrics
ws_inv = wb.create_sheet("Inventory_Metrics")
ws_inv.views.sheetView[0].showGridLines = True

inv_headers = ["Snapshot_Date", "Warehouse_Location", "Category", "Current_Stock_Qty", "Safety_Stock_Threshold", "Stock_Compliance_Pct", "Supplier_Lead_Time_Days", "Status"]
ws_inv.append(inv_headers)

for col_idx in range(1, len(inv_headers) + 1):
    cell = ws_inv.cell(row=1, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

warehouses = ["Denver Warehouse", "Seattle Import Hub", "Chicago Depot"]
# Probabilities of safety stock compliance based on dashboard (Denver 98%, Seattle 76%, Chicago 42%)
wh_compliance_targets = {
    "Denver Warehouse": 0.98,
    "Seattle Import Hub": 0.76,
    "Chicago Depot": 0.42
}

wh_lead_times = {
    "Denver Warehouse": (2, 5),
    "Seattle Import Hub": (7, 15),
    "Chicago Depot": (4, 10)
}

num_inv_rows = 300
for i in range(num_inv_rows):
    date_val = start_date + datetime.timedelta(days=random.randint(0, 240))
    wh = random.choice(warehouses)
    category = random.choice(categories)
    
    # Generate compliance % around target
    target = wh_compliance_targets[wh]
    compliance = round(random.normalvariate(target, 0.08), 3)
    compliance = max(0.1, min(1.3, compliance)) # Bound between 10% and 130%
    
    safety_stock = random.choice([100, 250, 500, 1000])
    current_stock = int(safety_stock * compliance)
    
    lead_time = random.randint(*wh_lead_times[wh])
    
    # Status
    if compliance < 0.5:
        status = "Critical Low"
    elif compliance < 0.8:
        status = "Warning"
    else:
        status = "Normal"
        
    ws_inv.append([date_val, wh, category, current_stock, safety_stock, compliance, lead_time, status])

# Style Inventory Data
for r_idx in range(2, num_inv_rows + 2):
    is_even = (r_idx % 2 == 0)
    for c_idx in range(1, len(inv_headers) + 1):
        cell = ws_inv.cell(row=r_idx, column=c_idx)
        cell.font = font_regular
        cell.border = thin_border
        if is_even:
            cell.fill = fill_zebra
            
        # Alignments & formats
        if c_idx == 1:
            cell.alignment = align_center
            cell.number_format = 'yyyy-mm-dd'
        elif c_idx in [2, 3]:
            cell.alignment = align_left
        elif c_idx in [4, 5, 7]:
            cell.alignment = align_right
            cell.number_format = '#,##0'
        elif c_idx == 6:
            cell.alignment = align_right
            cell.number_format = '0.0%'
        elif c_idx == 8:
            cell.alignment = align_center
            if cell.value == "Critical Low":
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # light red
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
            elif cell.value == "Warning":
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # light amber
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="92400E")
            elif cell.value == "Normal" and is_even:
                cell.fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid") # light green
                cell.font = Font(name="Segoe UI", size=10, color="065F46")
            elif cell.value == "Normal":
                cell.font = Font(name="Segoe UI", size=10, color="065F46")


# Sheet 4: KPI_Summary (Dashboard Formulas)
ws_kpi = wb.create_sheet("KPI_Summary")
ws_kpi.views.sheetView[0].showGridLines = True

# Title
ws_kpi.cell(row=1, column=1, value="Operational Dashboard Summary").font = Font(name="Segoe UI", size=14, bold=True, color="4F46E5")
ws_kpi.cell(row=2, column=1, value="Calculated using Excel Formulas from raw databases").font = Font(name="Segoe UI", size=9, italic=True)

# Table Header
kpi_headers = ["Metric Title", "Current Value", "Excel Formula Used", "Target Benchmark"]
ws_kpi.append([]) # spacing row
ws_kpi.append(kpi_headers)

header_row = 4
for col_idx in range(1, len(kpi_headers) + 1):
    cell = ws_kpi.cell(row=header_row, column=col_idx)
    cell.font = font_header
    cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
    cell.alignment = align_center

# Add KPIs with Formulas
kpis_data = [
    [
        "Total Tracked Revenue", 
        f"=SUM(Sales_Transactions!D2:D{num_sales_rows + 1})", 
        f"SUM(Sales_Transactions!D2:D{num_sales_rows + 1})", 
        "$1,000,000"
    ],
    [
        "Operations Leakage (Value)", 
        f"=SUM(Sales_Transactions!F2:F{num_sales_rows + 1})", 
        f"SUM(Sales_Transactions!F2:F{num_sales_rows + 1})", 
        "Under 5% of Revenue"
    ],
    [
        "Overall Safety Stock Compliance", 
        f"=AVERAGE(Inventory_Metrics!F2:F{num_inv_rows + 1})", 
        f"AVERAGE(Inventory_Metrics!F2:F{num_inv_rows + 1})", 
        "95.0%"
    ],
    [
        "Process Error Rate", 
        f'=COUNTIF(Sales_Transactions!H2:H{num_sales_rows + 1},"Yes")/COUNTA(Sales_Transactions!H2:H{num_sales_rows + 1})', 
        f'COUNTIF(Sales_Transactions!H2:H{num_sales_rows + 1}, "Yes") / COUNTA(...)', 
        "Under 2.5%"
    ]
]

for idx, row in enumerate(kpis_data, 1):
    r_idx = header_row + idx
    ws_kpi.append(row)
    
    # Format cells
    c_title = ws_kpi.cell(row=r_idx, column=1)
    c_title.font = font_bold
    c_title.alignment = align_left
    c_title.border = thin_border
    
    c_val = ws_kpi.cell(row=r_idx, column=2)
    c_val.font = font_bold
    c_val.alignment = align_right
    c_val.border = thin_border
    
    # Formats for value
    if idx == 1 or idx == 2:
        c_val.number_format = '$#,##0.00'
    elif idx == 3 or idx == 4:
        c_val.number_format = '0.0%'
        
    c_formula = ws_kpi.cell(row=r_idx, column=3)
    c_formula.font = font_regular
    c_formula.alignment = align_left
    c_formula.border = thin_border
    
    c_bench = ws_kpi.cell(row=r_idx, column=4)
    c_bench.font = font_regular
    c_bench.alignment = align_center
    c_bench.border = thin_border

# Add Department breakdown table below
ws_kpi.append([])
ws_kpi.append([])

breakdown_start_row = header_row + len(kpis_data) + 3
ws_kpi.cell(row=breakdown_start_row, column=1, value="Department Breakdown Summary").font = Font(name="Segoe UI", size=11, bold=True, color="1F2937")

breakdown_headers = ["Category", "Total Revenue", "Leakage Value", "Error Rate"]
ws_kpi.append(breakdown_headers)

for col_idx in range(1, len(breakdown_headers) + 1):
    cell = ws_kpi.cell(row=breakdown_start_row + 1, column=col_idx)
    cell.font = font_header
    cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid") # Charcoal
    cell.alignment = align_center

for idx, cat in enumerate(categories, 1):
    row_num = breakdown_start_row + 1 + idx
    # Formulas for Category breakdown
    rev_formula = f'=SUMIF(Sales_Transactions!C2:C{num_sales_rows + 1}, "{cat}", Sales_Transactions!D2:D{num_sales_rows + 1})'
    leak_formula = f'=SUMIF(Sales_Transactions!C2:C{num_sales_rows + 1}, "{cat}", Sales_Transactions!F2:F{num_sales_rows + 1})'
    err_formula = f'=COUNTIFS(Sales_Transactions!C2:C{num_sales_rows + 1}, "{cat}", Sales_Transactions!H2:H{num_sales_rows + 1}, "Yes")/COUNTIF(Sales_Transactions!C2:C{num_sales_rows + 1}, "{cat}")'
    
    ws_kpi.append([cat, rev_formula, leak_formula, err_formula])
    
    # Styles
    for col_idx in range(1, len(breakdown_headers) + 1):
        cell = ws_kpi.cell(row=row_num, column=col_idx)
        cell.font = font_regular
        cell.border = thin_border
        if col_idx == 1:
            cell.font = font_bold
            cell.alignment = align_left
        elif col_idx in [2, 3]:
            cell.alignment = align_right
            cell.number_format = '$#,##0.00'
        elif col_idx == 4:
            cell.alignment = align_right
            cell.number_format = '0.0%'

# Auto fit column widths for all sheets
for sheet in wb.worksheets:
    if sheet.title == "README":
        continue
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Calculate max length
        for cell in col:
            val_str = str(cell.value or '')
            if val_str.startswith('='):
                val_str = "Formula_Placeholder_String"
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save workbook
wb.save(excel_path)
print(f"Excel workbook generated successfully at: {excel_path}")
